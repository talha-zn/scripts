import json
import os
import pathlib

# specify the path to the directory containing JSON files
path_to_input_fils  = "./mappings"
input_files = os.listdir(path_to_input_fils)

def read_json(file_path):
    try:
        with open(file_path, 'r') as file:
            data = json.load(file)
            return data
    except FileNotFoundError:
        print(f"ERROR - File not found: {file_path}")
        return None
    except json.JSONDecodeError:
        print(f"ERROR - Error decoding JSON from file: {file_path}")
        return None

def extract_input_file_type(filename):
    name, _ = os.path.splitext(filename)
    return name.split('_')[-1] 

def extract_mapper_fields(mapper_fields):
    fields = {}
    for key, value in mapper_fields.items():
        fields[key] = {
            "type": value.get("type", ""),
            "source": value.get("source", "")
        }
    return fields 

def create_file_name(filetype):
    return "output.xlsx"

def create_excel_sheet(fields, filename, file_label, sheet_name):
    import pandas as pd
    from openpyxl import load_workbook
    print("INFO - Processing", file_label)

    df_new = pd.DataFrame([
        {"Type": v["type"], "Field Name": k, file_label: v["source"]}
        for k, v in fields.items()
    ])

    try:
        if pathlib.Path(filename).exists():
            # Check if the sheet exists
            wb = load_workbook(filename)
            if sheet_name.islower():
                print(f"ERROR - data type cannot be in lowercase: {file_label}")
                raise SystemExit(1)
            if sheet_name in wb.sheetnames:
                df_existing = pd.read_excel(filename, sheet_name=sheet_name)
                # Check for type mismatch
                if not df_existing.empty and not df_new.empty:
                    for _, row in df_new.iterrows():
                        field_name = row["Field Name"]
                        new_type = row["Type"]
                        existing_row = df_existing[df_existing["Field Name"] == field_name]
                        if not existing_row.empty:
                            existing_type = existing_row.iloc[0]["Type"]
                            if existing_type != new_type:
                                print(f"ERROR - Type mismatch for field '{field_name}': existing type '{existing_type}', new type '{new_type}'in file '{file_label}'")
                                print("ERROR - Fix type mismatch before merging.")
                                print(filename)
                                os.remove(filename) 
                                raise SystemExit(1)
                df_merged = pd.merge(df_existing, df_new[["Type", "Field Name", file_label]], on="Field Name", how="outer")
                df_merged["Type"] = df_merged["Type_x"].combine_first(df_merged["Type_y"])
                df_merged = df_merged.drop(columns=["Type_x", "Type_y"])
                cols = ["Type", "Field Name"] + [c for c in df_merged.columns if c not in ["Type", "Field Name"]]
                df_merged = df_merged[cols]
            else:
                df_merged = df_new

            # Write to the sheet (overwrite or create)
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_merged.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"INFO - Updated/created sheet '{sheet_name}' in Excel file: {filename}")
        else:
            df_new.to_excel(filename, sheet_name=sheet_name, index=False)
            print(f"INFO - Excel file created: {filename}")
    except Exception as e:
        # Catch-all in case delete fails
        print(f"ERROR - Aborting due to error: {e}")
        if pathlib.Path(filename).exists():
            os.remove(filename)
        raise  

if __name__ == "__main__":
    for file in input_files:
        filetype = extract_input_file_type(file)
        FILE_PATH = os.path.join(path_to_input_fils, file)
        data = read_json(FILE_PATH)
        if data:
            fields = extract_mapper_fields(data["fields"])
            file_label = os.path.splitext(file)[0] 
            sheet_name = file_label.split('_')[-1]
            create_excel_sheet(fields, create_file_name(filetype), file_label, sheet_name)