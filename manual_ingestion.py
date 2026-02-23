import openpyxl
from typing import Any
from clickhouse_driver import Client as ClickhouseClient
from datetime import datetime, timezone
import uuid


def establish_db_connection():
    """
    Establish and return a ClickHouse client.
    Use native TCP port (9000) for clickhouse_driver.
    """
    # Filling in some defaults based on usual patterns if user didn't specify,
    # but I'll leave them as they were mostly, just ensuring it's solid.
    host = "localhost"
    port = 9000
    user = "admin"
    password = "SecureAdminPass123!"
    database = "cvl_test_new"

    client = ClickhouseClient(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        settings={"connect_timeout": 10},
    )

    # test connection
    client.execute("SELECT 1")
    return client


def process_existing_data(ch_client):
    """
    Mark existing manual_ingestion rows as not latest
    """
    query = """
    ALTER TABLE ZCN_DEVICES UPDATE ZCN_IS_LATEST = 0
    WHERE ZCN_IS_LATEST = 1 AND ZCN_SOURCE = 'manual_ingestion' AND ZCN_CATEGORY = 'cmdb_manual_ingestion'
    """
    try:
        ch_client.execute(query)
        print("✓ Existing data marked as not latest")
    except Exception as e:
        print(f"Update error: {e}")
        raise


def get_column_mapping(ch_client, excel_headers):
    """
    Check ClickHouse for column existence and return a mapping.
    If column doesn't exist, append __. If still doesn't exist, add it.
    """
    # Get existing columns
    result = ch_client.execute("DESCRIBE TABLE ZCN_DEVICES")
    existing_columns = {col[0].upper() for col in result}

    mapping = {}
    for header in excel_headers:
        upper_header = header.upper().replace(" ", "_")

        # Check direct match
        if upper_header in existing_columns:
            mapping[header] = upper_header
            continue

        # Check with __ suffix
        custom_col = f"{upper_header}__"
        if custom_col in existing_columns:
            mapping[header] = custom_col
            continue

        # If not present at all, add it as a custom column
        print(f"Column {upper_header} not found, adding as {custom_col}")
        alter_query = (
            f"ALTER TABLE ZCN_DEVICES ADD COLUMN {custom_col} Nullable(String)"
        )
        try:
            ch_client.execute(alter_query)
            mapping[header] = custom_col
        except Exception as e:
            print(f"✗ Failed to add column {custom_col}: {e}")
            # Fallback to a generic string or skip? Usually we want to succeed.
            mapping[header] = custom_col

    return mapping


def generate_collection_id():
    """Generate collection ID in format: cvlYEARMONTHDATEHOURMINUTE"""
    now = datetime.now(timezone.utc)
    return f"cvl{now.strftime('%Y%m%d%H%M')}"


def read_excel(file_path):
    """Read Excel file and return metadata and data rows"""
    book = openpyxl.load_workbook(file_path)
    ws = book.active

    headers = [cell.value for cell in ws[1]]
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_rows.append(row)

    return headers, data_rows


def sanitize_cell(cell):
    if cell is None:
        return ""
    elif isinstance(cell, bool):
        return str(cell)
    elif isinstance(cell, (int, float, str, datetime)):
        return cell
    else:
        return str(cell)


def dump_data_batch(ch_client, headers, rows, mapping, collection_id):
    """
    Insert a batch of rows into ClickHouse with dynamic columns
    """
    now = datetime.now(timezone.utc)
    zcn_collected_at_naive = now.replace(tzinfo=None)
    zcn_source = "manual_ingestion"
    zcn_category = "cmdb_manual_ingestion"
    zcn_is_latest = 1

    # Prepare final columns list
    ch_columns = [mapping[h] for h in headers]

    # Metadata columns
    metadata_cols = [
        "ZCN_SOURCE",
        "ZCN_ID",
        "ZCN_COLLECTED_AT",
        "ZCN_CATEGORY",
        "ZCN_IS_LATEST",
        "ZCN_COLLECTION_ID",
    ]

    insert_columns = ch_columns + metadata_cols

    final_rows = []
    for row in rows:
        # Sanitize data columns
        safe_data = [sanitize_cell(cell) for cell in row]

        # Add metadata
        zcn_id = str(uuid.uuid4())
        metadata_values = [
            zcn_source,
            zcn_id,
            zcn_collected_at_naive,
            zcn_category,
            zcn_is_latest,
            collection_id,
        ]

        final_rows.append(tuple(safe_data + metadata_values))

    # Construct dynamic query
    cols_str = ", ".join(insert_columns)
    query = f"INSERT INTO ZCN_DEVICES ({cols_str}) VALUES"

    try:
        ch_client.execute(query, final_rows)
        print(f"✓ Inserted {len(final_rows)} rows into columns: {cols_str}")
    except Exception as e:
        print(f"Insert error: {e}")
        raise


def process_excel(file_path, ch_client):
    """Process the Excel file and perform batch insertion"""
    headers, rows = read_excel(file_path)
    if not rows:
        print("No data found in Excel")
        return

    collection_id = generate_collection_id()

    # Dynamically get mapping and ensure columns exist
    mapping = get_column_mapping(ch_client, headers)

    # Perform batch insertion
    dump_data_batch(ch_client, headers, rows, mapping, collection_id)


if __name__ == "__main__":
    excel_file_path = "SEBI_assets_dummy.xlsx"
    try:
        ch_client = establish_db_connection()
        print("✓ Connected to ClickHouse")

        # mark old rows as not latest
        process_existing_data(ch_client)

        # Process Excel and batch insert
        process_excel(excel_file_path, ch_client)

    except Exception as e:
        print(f"✗ Error: {e}")
    finally:
        try:
            if "ch_client" in locals() and ch_client:
                ch_client.disconnect()
                print("✓ ClickHouse connection closed")
        except Exception:
            pass
