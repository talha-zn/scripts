import openpyxl
from typing import Any
from clickhouse_driver import Client as ClickhouseClient
from datetime import datetime, timezone
import uuid


def establish_db_connection():
    """
    Establish and return a ClickHouse client.
    Use native TCP port (9000) for clickhouse_driver.
    """
    # Filling in some defaults based on usual patterns if user didn't specify,
    # but I'll leave them as they were mostly, just ensuring it's solid.
    host = "localhost"
    port = 9000
    user = "admin"
    password = "SecureAdminPass123!"
    database = "cvl_test_new"

    client = ClickhouseClient(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        settings={"connect_timeout": 10},
    )

    # test connection
    client.execute("SELECT 1")
    return client


def process_existing_data(ch_client):
    """
    Mark existing manual_ingestion rows as not latest
    """
    query = """
    ALTER TABLE ZCN_DEVICES UPDATE ZCN_IS_LATEST = 0
    WHERE ZCN_IS_LATEST = 1 AND ZCN_SOURCE = 'manual_ingestion' AND ZCN_CATEGORY = 'cvl_inventory'
    """
    try:
        ch_client.execute(query)
        print("✓ Existing data marked as not latest")
    except Exception as e:
        print(f"Update error: {e}")
        raise


def sanitize_column_name(raw_header) -> str:
    """
    Convert an Excel header into a valid ClickHouse identifier:
      - coerce None / non-strings to string first
      - uppercase
      - replace spaces and any non-alphanumeric characters with '_'
      - collapse consecutive underscores into a single one
      - strip leading/trailing underscores
    Returns an empty string for headers that are blank or None.
    """
    import re
    if raw_header is None:
        return ""
    upper = str(raw_header).upper()
    safe = re.sub(r'[^A-Z0-9]+', '_', upper)   # replace bad chars
    safe = safe.strip('_')                        # trim edges
    return safe


def get_column_mapping(ch_client, excel_headers, sample_rows=None):
    """
    Check ClickHouse for column existence and return a mapping.
    If column doesn't exist in its sanitized form, append __. If still
    doesn't exist, ADD it as Nullable(DateTime) when the column contains
    datetime values, otherwise Nullable(String).
    Column names in SQL are backtick-quoted to avoid syntax errors.
    """
    result = ch_client.execute("DESCRIBE TABLE ZCN_DEVICES")
    # col[0]=name, col[1]=type string
    existing_col_types = {col[0].upper(): col[1] for col in result}
    existing_columns = set(existing_col_types.keys())

    def infer_col_type(col_idx):
        """Return ClickHouse type string based on first non-None sample value."""
        if sample_rows:
            for row in sample_rows:
                if col_idx < len(row) and row[col_idx] is not None and not isinstance(row[col_idx], bool):
                    if isinstance(row[col_idx], datetime):
                        return "Nullable(DateTime)"
                    break
        return "Nullable(String)"

    mapping = {}    # header -> ch column name
    col_types = {}  # header -> ClickHouse type string
    for idx, header in enumerate(excel_headers):
        safe_header = sanitize_column_name(header)

        # Skip completely blank / None header columns
        if not safe_header:
            print(f"⚠ Skipping blank/None header at column index {idx}")
            continue
        if safe_header in existing_columns:
            mapping[header] = safe_header
            col_types[header] = existing_col_types[safe_header]
            continue

        # Check with __ suffix (custom column convention)
        custom_col = f"{safe_header}__"
        if custom_col in existing_columns:
            mapping[header] = custom_col
            col_types[header] = existing_col_types[custom_col]
            continue

        # Not present at all — add it as a custom column
        col_type = infer_col_type(idx)
        print(f"Column {safe_header} not found, adding as {custom_col} ({col_type})")
        alter_query = (
            f"ALTER TABLE ZCN_DEVICES ADD COLUMN `{custom_col}` {col_type}"
        )
        try:
            ch_client.execute(alter_query)
            mapping[header] = custom_col
            col_types[header] = col_type
        except Exception as e:
            print(f"✗ Failed to add column {custom_col}: {e}")
            mapping[header] = custom_col
            col_types[header] = "Nullable(String)"

    return mapping, col_types


def generate_collection_id():
    """Generate collection ID in format: cvlYEARMONTHDATEHOURMINUTE"""
    now = datetime.now(timezone.utc)
    return f"cvl{now.strftime('%Y%m%d%H%M')}"


def read_excel(file_path, sheet_names=None):
    """
    Read Excel file and yield (sheet_name, headers, data_rows) for each
    sheet that matches the requested sheet_names list.
    If sheet_names is None or empty, all sheets are processed.
    """
    book = openpyxl.load_workbook(file_path)

    # Determine which sheets to process
    if sheet_names:
        sheets_to_process = [
            name for name in sheet_names if name in book.sheetnames
        ]
        skipped = [name for name in sheet_names if name not in book.sheetnames]
        if skipped:
            print(f"⚠ Sheets not found in workbook (skipped): {skipped}")
    else:
        sheets_to_process = book.sheetnames

    for sheet_name in sheets_to_process:
        ws = book[sheet_name]
        headers = [cell.value for cell in ws[1]]
        data_rows = [
            row for row in ws.iter_rows(min_row=2, values_only=True)
            if any(cell is not None for cell in row)
        ]
        yield sheet_name, headers, data_rows


def sanitize_cell(cell):
    """Generic sanitize — used when column type is unknown (falls back to string for datetime)."""
    if cell is None:
        return ""
    elif isinstance(cell, bool):
        return str(cell)
    elif isinstance(cell, datetime):
        return cell.isoformat()
    elif isinstance(cell, (int, float, str)):
        return cell
    else:
        return str(cell)


def sanitize_cell_typed(cell, ch_type: str):
    """Sanitize a cell value according to the actual ClickHouse column type."""
    if cell is None:
        return ""
    elif isinstance(cell, bool):
        return str(cell)
    elif isinstance(cell, datetime):
        if "DateTime" in ch_type:
            # DateTime column: pass a timezone-naive datetime object
            return cell.replace(tzinfo=None)
        else:
            # String (or other) column: convert to ISO string
            return cell.isoformat()
    elif isinstance(cell, (int, float, str)):
        return cell
    else:
        return str(cell)


def dump_data_batch(ch_client, headers, rows, mapping, col_types, collection_id):
    """
    Insert a batch of rows into ClickHouse with dynamic columns.
    col_types maps each excel header to its ClickHouse column type so
    datetime cells are serialized correctly (object for DateTime columns,
    ISO string for String columns).
    """
    now = datetime.now(timezone.utc)
    zcn_collected_at_naive = now.replace(tzinfo=None)
    zcn_source = "manual_ingestion"
    zcn_category = "cvl_inventory"
    zcn_is_latest = 1

    # Build ordered list of (header, ch_col_name) only for headers that have a mapping
    mapped_headers = [h for h in headers if h in mapping]
    ch_columns = [mapping[h] for h in mapped_headers]

    # Metadata columns
    metadata_cols = [
        "ZCN_SOURCE",
        "ZCN_ID",
        "ZCN_COLLECTED_AT",
        "ZCN_CATEGORY",
        "ZCN_IS_LATEST",
        "ZCN_COLLECTION_ID",
    ]

    all_columns = ch_columns + metadata_cols
    insert_columns = all_columns          # kept for the print below
    # Backtick-quote every column name going into the SQL string
    quoted_columns = [f"`{c}`" for c in all_columns]

    # Pre-calculate indices for mapped headers to avoid .index() in the loop
    header_to_idx = {h: headers.index(h) for h in headers if h is not None}
    header_info = []
    for h in mapped_headers:
        idx = header_to_idx.get(h)
        if idx is not None:
            header_info.append((h, idx, col_types.get(h, "Nullable(String)")))

    batch_size = 50000
    total_rows = len(rows)
    inserted_count = 0

    for i in range(0, total_rows, batch_size):
        batch = rows[i : i + batch_size]
        final_rows = []
        for row in batch:
            # Sanitize each cell using its actual ClickHouse column type
            safe_data = [
                sanitize_cell_typed(row[idx], ctype)
                for _, idx, ctype in header_info
            ]

            # Add metadata
            zcn_id = str(uuid.uuid4())
            metadata_values = [
                zcn_source,
                zcn_id,
                zcn_collected_at_naive,
                zcn_category,
                zcn_is_latest,
                collection_id,
            ]
            final_rows.append(tuple(safe_data + metadata_values))

        # Construct dynamic query with backtick-quoted column names
        cols_str = ", ".join(quoted_columns)
        query = f"INSERT INTO ZCN_DEVICES ({cols_str}) VALUES"

        try:
            ch_client.execute(query, final_rows)
            inserted_count += len(final_rows)
            print(f"✓ [{inserted_count}/{total_rows}] Inserted batch of {len(final_rows)} rows")
        except Exception as e:
            print(f"Insert error at row {i}: {e}")
            raise

    print(f"✅ Successfully inserted {inserted_count} rows into columns: {', '.join(insert_columns)}")


def process_excel(file_path, ch_client, sheet_names=None):
    """Process specific sheets from the Excel file and perform batch insertion"""
    collection_id = generate_collection_id()

    for sheet_name, headers, rows in read_excel(file_path, sheet_names):
        if not rows:
            print(f"No data found in sheet '{sheet_name}', skipping")
            continue

        print(f"\n📄 Processing sheet: '{sheet_name}' ({len(rows)} rows)")

        # Dynamically get mapping + column types; ensure columns exist
        mapping, col_types = get_column_mapping(ch_client, headers, sample_rows=rows)

        # Perform batch insertion
        dump_data_batch(ch_client, headers, rows, mapping, col_types, collection_id)


if __name__ == "__main__":
    excel_file_path = "asset.xlsx"

    # Specify which sheets to ingest. Set to None (or []) to process ALL sheets.
    target_sheets = ["CVL Inventory"]

    try:
        ch_client = establish_db_connection()
        print("✓ Connected to ClickHouse")

        # mark old rows as not latest
        process_existing_data(ch_client)

        # Process only the specified sheets and batch insert
        process_excel(excel_file_path, ch_client, sheet_names=target_sheets)

    except Exception as e:
        print(f"✗ Error: {e}")
    finally:
        try:
            if "ch_client" in locals() and ch_client:
                ch_client.disconnect()
                print("✓ ClickHouse connection closed")
        except Exception:
            pass
